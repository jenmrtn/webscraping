from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
from twilio.rest import Client
import keys

client= Client(keys.account_sid, keys.auth_token)
fun_font = Font(name='Calibri', size=15,  color='00008000')
goodfont= Font(color='0000FF00')
badfont=Font(color='00FF0000')
TWNumber= "+15732608058"

mphone= '+14178253408'

url = 'https://cryptoslate.com/coins/'
# Request in case 404 Forbidden error
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}

req = Request(url, headers=headers)

webpage= urlopen(req).read()

soup = BeautifulSoup(webpage, 'html.parser')

print(soup.title)

wb= xl.Workbook()

ws =wb.active

wb.title= "Top 5 CryptoCurrencies"

ws['A1']='Rank'
ws['B1']= 'Name and Symbol'
ws['C1']= 'Price'
ws['D1']= '% Change'
ws['E1']= 'Corresponding Price'

table_rows= soup.findAll("tr")

for i in range(1,6):
    td = table_rows[i].findAll('td')
    rank= td[0].text
    xlprice= td[2].text
    xlpercent= td[3].text
    price = float(td[2].text.replace('$','').replace(' ','').replace(',',''))
    percentage = float(td[3].text.replace('%','').replace('-',''))
    name=td[1].text
    textname=td[1].text.replace(' ','')
    if '-' in td[3].text:
        percentage=percentage* -1
        pricepercent= percentage/100
    else:
        pricepercent=percentage/100
    corprice=price/1+pricepercent
    xlnewprice = '${:,.2f}'.format(corprice)
    ws['A' +str(i+1)]= rank
    ws['B' +str(i+1)]= name
    ws['C' +str(i+1)]= xlprice
    ws['D' +str(i+1)]= xlpercent
    ws['E' +str(i+1)]= xlnewprice
    if corprice> price:
        ws['E' +str(i+1)].font= goodfont
    else:
        ws['E' +str(i+1)].font= badfont
    ws['A' +str(i+1)].font= fun_font

    if 'BTC' or 'ETH' in name:
        if corprice-price < -5 or corprice-price > 5:
            textmsg= client.messages.create(to=mphone, from_=TWNumber, body= f'Alert {name} has changed over $5')
header_font = Font(size=16, bold=True)


for cell in ws[1:1]:
    cell.font= header_font



ws.column_dimensions['A'].width= 15
ws.column_dimensions['B'].width= 25
ws.column_dimensions['C'].width= 25
ws.column_dimensions['D'].width= 25
ws.column_dimensions['E'].width= 25

wb.save('Top5CryptoCurrencies.xlsx')

client= Client(keys.account_sid, keys.auth_token)

TWNumber= "+15732608058"

mphone= '+14178253408'

